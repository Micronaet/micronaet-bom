# -*- coding: utf-8 -*-
###############################################################################
#
#    Copyright (C) 2001-2014 Micronaet SRL (<http://www.micronaet.it>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as published
#    by the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################
import os
import sys
import logging
import openerp
import openerp.netsvc as netsvc
import openerp.addons.decimal_precision as dp
from openerp.osv import fields, osv, expression, orm
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openerp import SUPERUSER_ID, api
from openerp import tools
from openerp.tools.translate import _
from openerp.tools.float_utils import float_round as round
from openerp.tools import (DEFAULT_SERVER_DATE_FORMAT,
    DEFAULT_SERVER_DATETIME_FORMAT,
    DATETIME_FORMATS_MAP,
    float_compare)

# Excel:
import openpyxl
from openpyxl.styles import Border, Alignment, Font, numbers



_logger = logging.getLogger(__name__)


class SaleOrder(orm.Model):
    """ Model name: SaleOrder
    """

    _inherit = 'sale.order'

    _columns = {
        'simulation': fields.boolean('Simulation'),
        }


class ProductProductInventoryCategory(orm.Model):
    """ Model name: ProductProductInventoryCategory
    """

    _inherit = 'product.product.inventory.category'

    _columns = {
        'not_in_report': fields.boolean('No report'),
        }


class MrpProduction(orm.Model):
    """ Model name: Extra function for utility
    """

    _inherit = 'mrp.production'

    def integrate_stock_total_page_to_excel(self, filename):
        """ Integrate stock page in component / textilene report
        """
        # ---------------------------------------------------------------------
        # Parameters:
        # ---------------------------------------------------------------------
        total_col = 24
        stock_name = 'Magazzino'
        excluded_sheet = [
            'Non usati',
            stock_name,
            ]

        # ---------------------------------------------------------------------
        # Style:
        # ---------------------------------------------------------------------
        format_euro = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

        font = Font(
            name='Verdana', size=10, bold=False, italic=False,
            vertAlign=None, underline='none', strike=False,
            color='FF000000')
        font_bold = Font(
            name='Verdana', size=10, bold=True, italic=False,
            vertAlign=None, underline='none', strike=False,
            color='FF000000')

        alignment = Alignment(
            horizontal='general', vertical='bottom', text_rotation=0,
            wrap_text=False, shrink_to_fit=False, indent=0)
        alignment_center = Alignment(
            horizontal='center', vertical='bottom', text_rotation=0,
            wrap_text=False, shrink_to_fit=False, indent=0)

        # ---------------------------------------------------------------------
        #                         Open Excel file:
        # ---------------------------------------------------------------------
        filename = os.path.expanduser(filename)
        wb = openpyxl.load_workbook(filename=filename)
        total_page = {}

        # ---------------------------------------------------------------------
        # Load data from all sheets:
        # ---------------------------------------------------------------------
        for ws in wb._sheets:
            sheet_name = ws.title
            if sheet_name in excluded_sheet:  # Not Stock sheet
                continue

            total_page[sheet_name] = 0.0
            for row in range(1, ws.max_row + 1):
                partial = ws.cell(row=row, column=total_col).value

                # Keep only float:
                try:
                    partial = float(partial)
                except:
                    continue
                total_page[sheet_name] += partial

        try:
            ws = wb[stock_name]  # todo delete and recreate?
        except:
            ws = wb.create_sheet(stock_name, 0)  # Last page

        stock_total = 0.0

        # Setup columns:
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        # Write data for stock:
        row = 1
        label_cell = ws.cell(row=row, column=1)
        data_cell = ws.cell(row=row, column=2)

        # ---------------------------------------------------------------------
        # Header:
        # ---------------------------------------------------------------------
        label_cell.font = font_bold
        data_cell.font = font_bold
        label_cell.alignment = alignment_center
        data_cell.alignment = alignment_center

        label_cell.value = 'Categoria'
        data_cell.value = 'Valore'
        for category in sorted(total_page):
            row += 1
            label_cell = ws.cell(row=row, column=1)
            data_cell = ws.cell(row=row, column=2)

            # Style:
            label_cell.font = font_bold
            data_cell.font = font
            label_cell.alignment = alignment
            data_cell.alignment = alignment
            data_cell.number_format = format_euro

            # Data:
            label_cell.value = category
            data_cell.value = total_page[category]

            stock_total += total_page[category]

        # ---------------------------------------------------------------------
        # Last line with total stock:
        # ---------------------------------------------------------------------
        label_cell = ws.cell(row=row + 1, column=1)
        data_cell = ws.cell(row=row + 1, column=2)

        # Style:
        label_cell.font = font_bold
        data_cell.font = font_bold
        label_cell.alignment = alignment
        data_cell.alignment = alignment
        data_cell.number_format = format_euro

        label_cell.value = 'Totale magazzino'
        data_cell.value = stock_total

        wb.save(filename)
        return True
